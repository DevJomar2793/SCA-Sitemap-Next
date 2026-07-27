import re
from dataclasses import dataclass
from io import BytesIO
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

DEFAULT_VALUE = "Not provided"
MAX_IMPORTED_ROWS = 10_000
HEADER_SEARCH_ROWS = 25

HEADER_ALIASES = {
    "alpha": {"alpha"},
    "screen_number": {"screen", "screennumber"},
    "screen_type": {"screentype"},
    "screen_description": {"screendescription"},
    "file_label": {"filelabel"},
    "screen_label": {"screenlabel"},
    "notes": {"notes", "changeorder", "noteschangeorder"},
    "page_location": {
        "navigationinstructions",
        "pagelocation",
        "linkurl",
        "linkurls",
        "linksurls",
    },
}


class WorkbookImportError(ValueError):
    """Raised when an uploaded workbook cannot be safely imported."""


@dataclass(slots=True)
class WorkbookImportResult:
    records: list[dict[str, str]]
    skipped_count: int
    worksheet_count: int
    ignored_worksheets: list[str]


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", cell_text(value).lower())


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def find_header(
    worksheet: Any,
) -> tuple[int, dict[str, int]] | None:
    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=1,
            max_row=min(HEADER_SEARCH_ROWS, worksheet.max_row),
            values_only=True,
        ),
        start=1,
    ):
        columns: dict[str, int] = {}
        for column_index, value in enumerate(row):
            normalized = normalize_header(value)
            for field, aliases in HEADER_ALIASES.items():
                if normalized in aliases and field not in columns:
                    columns[field] = column_index

        if (
            {"alpha", "screen_number"}.issubset(columns)
            and columns["alpha"] == 0
            and columns["screen_number"] == 1
        ):
            return row_number, columns

    return None


def infer_identity(
    raw_alpha: str,
    raw_screen_number: str,
    worksheet_name: str,
) -> tuple[str, str]:
    if raw_alpha:
        return raw_alpha, raw_screen_number

    prefixed_number = re.fullmatch(
        r"([A-Za-z][A-Za-z-]*)[\s_-]*(\d+(?:\.\d+)?)",
        raw_screen_number,
    )
    if prefixed_number:
        return (
            prefixed_number.group(1).rstrip("-").upper(),
            prefixed_number.group(2),
        )

    inferred_alpha = worksheet_name.strip() or DEFAULT_VALUE
    return inferred_alpha, raw_screen_number


def row_value(row: tuple[Any, ...], columns: dict[str, int], field: str) -> str:
    column_index = columns.get(field)
    if column_index is None or column_index >= len(row):
        return ""
    return cell_text(row[column_index])


def parse_sitemap_workbook(contents: bytes) -> WorkbookImportResult:
    try:
        workbook = load_workbook(
            filename=BytesIO(contents),
            read_only=True,
            data_only=True,
        )
    except (BadZipFile, InvalidFileException, KeyError, OSError, ValueError) as error:
        raise WorkbookImportError("The uploaded file is not a valid .xlsx workbook") from error

    records: list[dict[str, str]] = []
    skipped_count = 0
    worksheet_count = 0
    ignored_worksheets: list[str] = []

    try:
        for worksheet in workbook.worksheets:
            header = find_header(worksheet)
            if header is None:
                ignored_worksheets.append(worksheet.title)
                continue

            worksheet_count += 1
            header_row, columns = header
            for row in worksheet.iter_rows(
                min_row=header_row + 1,
                values_only=True,
            ):
                raw_values = {
                    field: row_value(row, columns, field)
                    for field in HEADER_ALIASES
                }
                has_content = any(raw_values.values())
                if not has_content:
                    continue

                has_screen_content = any(
                    raw_values[field]
                    for field in (
                        "screen_description",
                        "file_label",
                        "screen_label",
                    )
                )
                if not raw_values["screen_number"] or not has_screen_content:
                    skipped_count += 1
                    continue

                if len(records) >= MAX_IMPORTED_ROWS:
                    raise WorkbookImportError(
                        f"The workbook exceeds the {MAX_IMPORTED_ROWS:,}-row import limit"
                    )

                alpha, screen_number = infer_identity(
                    raw_values["alpha"],
                    raw_values["screen_number"],
                    worksheet.title,
                )
                records.append(
                    {
                        "alpha": alpha or DEFAULT_VALUE,
                        "screen_number": screen_number,
                        "screen_type": raw_values["screen_type"] or DEFAULT_VALUE,
                        "screen_description": (
                            raw_values["screen_description"] or DEFAULT_VALUE
                        ),
                        "file_label": raw_values["file_label"] or DEFAULT_VALUE,
                        "screen_label": raw_values["screen_label"] or DEFAULT_VALUE,
                        "notes": raw_values["notes"] or DEFAULT_VALUE,
                        "page_location": (
                            raw_values["page_location"] or DEFAULT_VALUE
                        ),
                    }
                )
    finally:
        workbook.close()

    if worksheet_count == 0:
        raise WorkbookImportError(
            "No worksheet contains the required Alpha and Screen# columns"
        )
    if not records:
        raise WorkbookImportError("No sitemap screen rows were found in the workbook")

    return WorkbookImportResult(
        records=records,
        skipped_count=skipped_count,
        worksheet_count=worksheet_count,
        ignored_worksheets=ignored_worksheets,
    )
