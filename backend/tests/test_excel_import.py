from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

import app.services.excel_import as excel_import
from app.api.routes.admin_pages import MAX_UPLOAD_BYTES

IMPORT_PATH = "/api/v1/import-sitemap-pages"
LIST_PATH = "/api/v1/get-admin-pages"
EXCEL_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
EXISTING_PAYLOAD = {
    "alpha": "A",
    "screen_number": "001",
    "screen_type": "Landing",
    "screen_description": "Main landing screen",
    "file_label": "landing.tsx",
    "screen_label": "Landing page",
    "notes": "Initial version",
    "page_location": "/",
}


def workbook_bytes(*, row_count: int = 2) -> bytes:
    workbook = Workbook()
    detail = workbook.active
    detail.title = "Detail"
    detail.append(
        [
            "alpha",
            "Screen#",
            "Screen Type",
            "Screen Description",
            "File Label",
            "Screen Label",
            "NOTES/Change Order",
            "Navigation Instructions",
        ]
    )
    detail.append(["Section heading"])
    rows = [
        [
            "A",
            3,
            None,
            "Dashboard",
            "A-03 Dashboard",
            "A-03",
            None,
            "Menu → Dashboard",
        ],
        [
            None,
            "MG-04",
            "List",
            "Manager list",
            None,
            "MG-04",
            "Ready",
            None,
        ],
    ]
    for row in rows[:row_count]:
        detail.append(row)

    summary = workbook.create_sheet("Summary")
    summary.append([None, "alpha", "numeric", "Description"])
    summary.append([None, "A", 3, "Duplicate dashboard"])

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def upload(
    client: TestClient,
    contents: bytes,
    *,
    filename: str = "screens.xlsx",
):
    return client.post(
        IMPORT_PATH,
        files={"file": (filename, contents, EXCEL_CONTENT_TYPE)},
    )


def test_import_replaces_existing_pages_and_reports_summary(
    client: TestClient,
) -> None:
    assert (
        client.post("/api/v1/add-admin-page", json=EXISTING_PAYLOAD).status_code
        == 201
    )

    response = upload(client, workbook_bytes())

    assert response.status_code == 200
    assert response.json() == {
        "imported_count": 2,
        "skipped_count": 1,
        "worksheet_count": 1,
        "ignored_worksheets": ["Summary"],
    }

    records = client.get(LIST_PATH).json()
    assert len(records) == 2
    assert records[0]["alpha"] == "A"
    assert records[0]["screen_number"] == "3"
    assert records[0]["screen_type"] == "Not provided"
    assert records[0]["notes"] == "Not provided"
    assert records[1]["alpha"] == "MG"
    assert records[1]["screen_number"] == "04"
    assert records[1]["page_location"] == "Not provided"


def test_reimport_replaces_instead_of_duplicating(client: TestClient) -> None:
    contents = workbook_bytes()

    assert upload(client, contents).status_code == 200
    assert upload(client, contents).status_code == 200

    assert len(client.get(LIST_PATH).json()) == 2


def test_invalid_extension_is_rejected(client: TestClient) -> None:
    response = upload(client, workbook_bytes(), filename="screens.xls")

    assert response.status_code == 400
    assert response.json()["detail"] == "Only .xlsx Excel files are supported"


def test_oversized_file_is_rejected(client: TestClient) -> None:
    response = upload(client, b"x" * (MAX_UPLOAD_BYTES + 1))

    assert response.status_code == 413


@pytest.mark.parametrize(
    ("contents", "expected_detail"),
    [
        (b"not an Excel file", "not a valid .xlsx workbook"),
        (
            None,
            "No worksheet contains the required Alpha and Screen# columns",
        ),
    ],
)
def test_invalid_workbook_does_not_replace_existing_pages(
    client: TestClient,
    contents: bytes | None,
    expected_detail: str,
) -> None:
    assert (
        client.post("/api/v1/add-admin-page", json=EXISTING_PAYLOAD).status_code
        == 201
    )

    if contents is None:
        workbook = Workbook()
        workbook.active.append(["Unrelated", "Columns"])
        buffer = BytesIO()
        workbook.save(buffer)
        workbook.close()
        contents = buffer.getvalue()

    response = upload(client, contents)

    assert response.status_code == 422
    assert expected_detail in response.json()["detail"]
    assert client.get(LIST_PATH).json()[0]["screen_label"] == "Landing page"


def test_row_limit_failure_preserves_existing_pages(
    client: TestClient,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    assert (
        client.post("/api/v1/add-admin-page", json=EXISTING_PAYLOAD).status_code
        == 201
    )
    monkeypatch.setattr(excel_import, "MAX_IMPORTED_ROWS", 1)

    response = upload(client, workbook_bytes())

    assert response.status_code == 422
    assert "1-row import limit" in response.json()["detail"]
    records = client.get(LIST_PATH).json()
    assert len(records) == 1
    assert records[0]["screen_label"] == "Landing page"
